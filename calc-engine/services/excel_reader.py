import os
import threading
from pathlib import Path

from .logger import get_logger

log = get_logger("excel_reader")

EXCEL_PATH = os.environ.get("EXCEL_PATH", "/app/excel/solar_calc.xlsx")

_lock   = threading.Lock()
_cache  = None


# ── Hardcoded fallback tables ─────────────────────────────────────────
# These are used when the Excel file is not present.
# Keys: (panel_thickness, tile_thickness, consequence_class, terrain_category, wind_velocity)

FALLBACK_BALLAST = {
    # (panel_mm, tile_mm, cc, terrain, wind) → tiles_per_shelter
    (30, 40, "CC2", "II",  24): 2,
    (30, 40, "CC2", "II",  25): 3,
    (30, 40, "CC2", "II",  26): 3,
    (30, 40, "CC2", "III", 24): 2,
    (30, 40, "CC2", "III", 25): 2,
    (30, 40, "CC2", "III", 26): 3,
    (30, 40, "CC3", "II",  24): 3,
    (30, 40, "CC3", "II",  25): 3,
    (30, 40, "CC3", "II",  26): 4,
    (25, 40, "CC2", "II",  24): 2,
    (25, 40, "CC2", "II",  25): 2,
    (25, 40, "CC2", "II",  26): 3,
    (35, 40, "CC2", "II",  24): 3,
    (35, 40, "CC2", "II",  25): 3,
    (35, 40, "CC2", "II",  26): 4,
}

FALLBACK_TILE_WEIGHTS = {
    40: 18.0,
    45: 22.5,
    50: 27.0,
}

PRICING_OVERRIDES_PATH = os.environ.get(
    "PRICING_OVERRIDES_PATH",
    "/app/config/pricing_overrides.json"
)

DEFAULT_PRICES = {
    "200101": 185.0,
    "200201": 12.5,
    "200301": 12.5,
    "100701": 3.2,
    "100801": 3.2,
    "100901": 3.2,
    "100601": 8.9,
}

DEFAULT_TIERS = [
    {"min_qty": 500, "discount_pct": 15},
    {"min_qty": 200, "discount_pct": 10},
    {"min_qty": 100, "discount_pct":  5},
    {"min_qty":   0, "discount_pct":  0},
]

_pricing_cache = None


def _load_pricing_overrides():
    global _pricing_cache
    import json

    prices = dict(DEFAULT_PRICES)
    tiers  = list(DEFAULT_TIERS)

    p = Path(PRICING_OVERRIDES_PATH)
    if p.exists():
        try:
            data = json.loads(p.read_text())
            if "unit_prices"    in data: prices.update(data["unit_prices"])
            if "discount_tiers" in data: tiers = data["discount_tiers"]
            log.info("pricing_overrides_loaded", extra={"path": str(p)})
        except Exception as exc:
            log.warning("pricing_overrides_failed", extra={"error": str(exc)})

    _pricing_cache = {"unit_prices": prices, "discount_tiers": tiers}
    return _pricing_cache


def get_pricing():
    if _pricing_cache is None:
        return _load_pricing_overrides()
    return _pricing_cache


def reload_config():
    global _pricing_cache, _cache
    with _lock:
        _pricing_cache = None
        _cache         = None
    _load_pricing_overrides()
    log.info("config_reloaded")


def get_ballast_tiles(
    panel_thickness: int,
    tile_thickness:  int,
    consequence_class: str,
    terrain_category:  str,
    wind_velocity:     int,
) -> int:
    key = (panel_thickness, tile_thickness, consequence_class, terrain_category, wind_velocity)

    excel_path = Path(EXCEL_PATH)
    if excel_path.exists():
        try:
            return _read_from_excel(key)
        except Exception as exc:
            log.warning("excel_read_failed", extra={"error": str(exc), "key": str(key)})

    result = FALLBACK_BALLAST.get(key)
    if result is not None:
        return result

    # Nearest-neighbour fallback — find closest wind velocity
    matches = [
        (abs(k[4] - wind_velocity), v)
        for k, v in FALLBACK_BALLAST.items()
        if k[0] == panel_thickness
        and k[1] == tile_thickness
        and k[2] == consequence_class
    ]
    if matches:
        matches.sort(key=lambda x: x[0])
        return matches[0][1]

    log.warning("ballast_fallback_default", extra={"key": str(key)})
    return 2


def get_tile_weight(tile_thickness: int) -> float:
    return FALLBACK_TILE_WEIGHTS.get(tile_thickness, 18.0)


def _read_from_excel(key):
    import openpyxl
    with _lock:
        global _cache
        if _cache is None:
            wb     = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
            _cache = wb
        wb = _cache

    panel_thickness, tile_thickness, cc, terrain, wind = key

    sheet_name = f"CC_{cc}"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name} not found in Excel")

    ws = wb[sheet_name]
    for row in ws.iter_rows(values_only=True):
        if (
            row[0] == panel_thickness and
            row[1] == tile_thickness  and
            row[2] == terrain         and
            row[3] == wind
        ):
            return int(row[4])

    raise ValueError(f"No row found for key {key}")